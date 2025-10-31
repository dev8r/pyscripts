import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import load_workbook
import xlrd

class ExcelViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Workbook Viewer")
        
        # Dropdown for worksheet selection
        self.sheet_selector = ttk.Combobox(root, state="readonly")
        self.sheet_selector.pack(pady=10)
        self.sheet_selector.bind("<<ComboboxSelected>>", self.display_sheet)

        # Frame for table and scrollbars
        table_frame = tk.Canvas(root, width=800, height=400, scrollregion=canvas.bbox("all"))
        table_frame.pack(expand=True, fill="both", pady=10)

        # Table for displaying worksheet content
        self.table = ttk.Treeview(table_frame, show="headings")
        self.table.pack(side="left", expand=True, fill="both")

        # Scrollbars for the table
        x_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
        y_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
       # self.table.configure(xscrollcommand=x_scrollbar.set, yscrollcommand=y_scrollbar.set)
        x_scrollbar.pack(side="bottom", fill="x")
        y_scrollbar.pack(side="right", fill="y")

        # Button to load Excel file
        self.load_button = ttk.Button(root, text="Load Excel File", command=self.load_excel)
        self.load_button.pack(pady=10)

    def load_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return

        if file_path.endswith(".xls"):
            workbook = xlrd.open_workbook(file_path)
            self.workbook = {sheet.name: [[cell.value for cell in row] for row in sheet.get_rows()] for sheet in workbook.sheets()}
            self.sheet_selector["values"] = list(self.workbook.keys())
        else:
            self.workbook = load_workbook(file_path)
            self.sheet_selector["values"] = self.workbook.sheetnames

        if self.sheet_selector["values"]:
            self.sheet_selector.current(0)
            self.display_sheet()
        if not file_path:
            return

        self.workbook = load_workbook(file_path)
        self.sheet_selector["values"] = self.workbook.sheetnames
        if self.workbook.sheetnames:
            self.sheet_selector.current(0)
            self.display_sheet()

    def display_sheet(self, event=None):
        sheet_name = self.sheet_selector.get()
        if not sheet_name:
            return

        sheet = self.workbook[sheet_name]

        # Clear the table
        self.table.delete(*self.table.get_children())
        self.table["columns"] = []

        # Add columns and rows to the table
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index == 0:
                self.table["columns"] = [f"Column {i+1}" for i in range(len(row))]
                for col in self.table["columns"]:
                    self.table.heading(col, text=col)
                    self.table.column(col, width=100, minwidth=100)  # Set minimum column width
            self.table.insert("", "end", values=row)
            
root = tk.Tk()
app = ExcelViewer(root)
root.geometry("800x600")
root.mainloop()
